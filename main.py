import pdfplumber
import os
import openpyxl

# список пдф файлов с места запуска
list_pdf = []
for i, k, j in os.walk(os.getcwd()):
    for f in j:
        if f.split('.')[-1].lower() == 'pdf':
            list_pdf.append(f'{i}\\{f}')
count_file = len(list_pdf)
count_pages = 0
count_error_files = []
# создать если нет эксель файла
if not os.path.exists('out.xlsx'):
    wb = openpyxl.Workbook()
    sets = wb.worksheets[0]
    wb.save(filename='out.xlsx')

# получаем данные с каждой страницы
file_ex = 'out.xlsx'
book = openpyxl.load_workbook(file_ex)
for file in range(len(list_pdf)):

    # sheets = book.sheetnames  # все листы
    sheet = book.worksheets[0]  # для начала первый лист
    pdf = pdfplumber.open(list_pdf[file])

    try:
        for p in range(len(pdf.pages)):
            # print(f'Текущий документ {list_pdf[file]}, страница {p}')
            all_set_kbk = []
            page = pdf.pages[p]
            text = page.extract_text()
            list_text = text.split()
            kbk = list_text[list_text.index('поле') + 1]
            if kbk not in all_set_kbk:
                all_set_kbk.append(kbk)
            number = list_text[list_text.index('ПОРУЧЕНИЕ') + 2]
            date = list_text[list_text.index('ПОРУЧЕНИЕ') + 3]
            inn = list_text[list_text.index('прописью') + 2]
            kpp = list_text[list_text.index('прописью') + 4]
            oktmo = list_text[list_text.index('поле') + 2]
            amount = list_text[list_text.index('прописью') + 6]
            purpose = []
            plat = []

            for i in list_text[list_text.index('прописью') + 7:88]:
                if i == 'Сч.':
                    break
                if i == 'БИК':
                    continue
                # if i.isdigit():
                #     continue
                plat.append(i)
            plat = ' '.join(plat)

            for s in list_text[list_text.index('Получатель') + 11:]:
                if s == 'Назначение':
                    break
                purpose.append(s)
            purpose = " ".join(purpose)

            add_in_row = [number, date, inn, kpp, oktmo, kbk, amount, plat, purpose]

            sheet.append(add_in_row)
            count_pages += 1
    except ValueError:
        # print(f'Не обработанный файл — {list_pdf[file]}')
        count_error_files.append(list_pdf[file])
        continue

book.save(file_ex)
if len(count_error_files) > 0:
    print(
        f"Количество документов — {count_file}, количество страниц — {count_pages},"
        f" количество необработанных файлов {len(count_error_files)}\nНеобработанный файл:")
    for i in count_error_files:
        print(i)
else:
    print(f"Количество пдф документов — {count_file}, количество страниц — {count_pages}")
input("Нажмите Enter чтобы закрыть окно ")
